import os
import pandas as pd
from openpyxl import load_workbook

def save_to_excel(records, original_pdf_path):
    """
    Save the parsed data to an Excel file with specified column order.

    Args:
        records (list of dict): List of dictionaries containing the data to save.
        original_pdf_path (str): Path to the original PDF to derive the Excel file name.
    """
    if not records:
        print("No records to save to Excel.")
        return

    # Create a DataFrame from the records
    df = pd.DataFrame(records)

    # Define the desired column order, including Amount as column F
    desired_order = ["Date", "Company", "Location", "Purchase Order", "Description", "Amount"]

    # Reorder the DataFrame columns based on the desired order.
    # Columns not in desired_order will be dropped.
    df = df.reindex(columns=desired_order)

    # Define the Excel file path
    base_name = os.path.splitext(os.path.basename(original_pdf_path))[0]
    excel_file_path = os.path.join(os.path.dirname(original_pdf_path), f"{base_name}_extracted.xlsx")

    try:
        # Save the reordered DataFrame to an Excel file
        df.to_excel(excel_file_path, index=False)
        print(f"Extracted data saved to Excel: {excel_file_path}")
    except Exception as e:
        print(f"Failed to save data to Excel file: {e}")


def append_to_existing_excel(records, existing_excel_path):
    """
    Append parsed data to an existing Excel file, preserving the first 9 rows and headers on row 9.

    Args:
        records (list of dict): List of dictionaries containing the data to append.
        existing_excel_path (str): Path to the existing Excel file.
    """
    if not records:
        print("No records to append to Excel.")
        return

    # Load the existing workbook and select the active worksheet
    wb = load_workbook(existing_excel_path)
    ws = wb.active

    # Define the desired column order
    desired_order = ["Date", "Company", "Location", "Purchase Order", "Description", "Amount"]

    # Determine the starting row for new data:
    # Ensure we don't overwrite the first 9 rows.
    start_row = max(ws.max_row + 1, 10)  # Start at row 10 if file has fewer than 9 filled rows

    # Append records one by one in the desired order
    for record in records:
        row_values = [
            record.get("Date", "N/A"),
            record.get("Company", "N/A"),
            record.get("Location", "N/A"),
            record.get("Purchase Order", "N/A"),
            record.get("Description", "N/A"),
            record.get("Amount", "N/A")
        ]
        # Write values to the row starting at start_row
        for col_index, value in enumerate(row_values, start=1):  # Columns start at 1 (A)
            ws.cell(row=start_row, column=col_index, value=value)
        start_row += 1  # Move to next row for next record

    try:
        wb.save(existing_excel_path)
        print(f"Data appended to existing Excel: {existing_excel_path}")
    except Exception as e:
        print(f"Failed to save data to Excel file: {e}")
