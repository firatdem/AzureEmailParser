import pandas as pd
from openpyxl import load_workbook

def append_to_existing_excel(records, existing_excel_path):
    wb = load_workbook(existing_excel_path)
    ws = wb.active
    start_row = max(ws.max_row + 1, 10)

    for record in records:
        row_values = [
            record.get("Date", "N/A"),
            record.get("Amount", "N/A")
        ]
        for col_index, value in enumerate(row_values, start=1):
            ws.cell(row=start_row, column=col_index, value=value)
        start_row += 1

    wb.save(existing_excel_path)
